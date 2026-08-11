#!/usr/bin/env python3
"""The email and spreadsheet built from the harvested universe."""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.harvest import xlsx_report
from src.harvest.email_report import build_email, render_html, summarise

# Hand-made universe: one exact UK premium, one exact UK discount, one ASX
# published discount, and one gross-assets estimate that is numerically the
# widest of all — which is exactly why it must stay out of the widest table.
ROWS = [
    dict(code="CTY", exchange="LSE", name="City of London", currency="GBP",
         market_cap=3.06e9, nta_total=2.97e9, nta_basis="net_shareholders_funds",
         nta_per_share=5.7545, price=5.93, discount=0.031,
         discount_basis="price_over_nav_net"),
    dict(code="SMT", exchange="LSE", name="Scottish Mortgage", currency="GBP",
         market_cap=1.43e10, nta_total=1.51e10, nta_basis="net_shareholders_funds",
         nta_per_share=14.1178, price=13.305, discount=-0.058,
         discount_basis="price_over_nav_net"),
    dict(code="AFI", exchange="ASX", name="Australian Foundation", currency="AUD",
         market_cap=8.7e9, nta_total=None, nta_basis=None,
         nta_per_share=7.76, price=6.55, discount=-0.156,
         discount_basis="published"),
    dict(code="PSH", exchange="LSE", name="Pershing Square", currency="GBP",
         market_cap=6.6e9, nta_total=1.18e10, nta_basis="gross_assets",
         nta_per_share=None, price=None, discount=-0.439,
         discount_basis="mcap_over_gross_assets"),
]
DROPPED = [dict(code="Total", name="Total", reason="not a recognisable stock code",
                market_cap=None, nta_per_share=None)]


def test_the_widest_table_never_promotes_the_biased_basis():
    """PSH's -43.9% is the widest number in the set and it is an estimate
    inflated by gearing. A 'widest discounts' list that leads with the known
    bias would be the report teaching its reader the wrong thing."""
    s = summarise(ROWS)
    widest = [r["code"] for r in s["widest"]]
    assert "PSH" not in widest
    assert widest[0] == "AFI"                      # -15.6%, published
    assert s["n_exact"] == 3                       # CTY, SMT, AFI


def test_the_email_html_carries_the_actual_numbers():
    s = summarise(ROWS)
    doc = render_html(s)
    assert "+3.1%" in doc          # CTY premium, sign intact
    assert "-15.6%" in doc         # AFI published discount
    assert "price_over_nav_net" in doc and "mcap_over_gross_assets" in doc
    assert "CTY" in doc and "biased wide by gearing" in doc
    # Gmail strips <style> blocks; everything must be inline.
    assert "<style" not in doc


def test_the_spreadsheet_round_trips_values_and_formats(tmp_path):
    path = str(tmp_path / "u.xlsx")
    xlsx_report.build(ROWS, DROPPED, {
        "subtitle": "4 funds", "facts": [("Funds", 4)],
        "by_basis": [("price_over_nav_net", 2, -0.058, -0.0135, 0.031, 1)],
        "notes": ["note"],
    }, path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert wb.sheetnames == ["Summary", "Universe", "Dropped"]
    ws = wb["Universe"]
    hdr = [c.value for c in ws[1]]
    codes = {ws.cell(row=r, column=1).value for r in range(2, 6)}
    assert codes == {"CTY", "SMT", "AFI", "PSH"}
    d = hdr.index("discount") + 1
    row_cty = next(r for r in range(2, 6) if ws.cell(row=r, column=1).value == "CTY")
    cell = ws.cell(row=row_cty, column=d)
    # Stored as the fraction, shown as a percent: cell and CSV always agree.
    assert cell.value == 0.031
    assert "%" in cell.number_format
    assert wb["Dropped"].cell(row=2, column=3).value == "not a recognisable stock code"


def test_the_email_has_html_text_and_the_attachment(tmp_path):
    path = str(tmp_path / "u.xlsx")
    xlsx_report.build(ROWS, DROPPED, {
        "subtitle": "s", "facts": [], "by_basis": [], "notes": [],
    }, path)
    s = summarise(ROWS)
    msg = build_email(s, path, "sender@gmail.com", "owner@gmail.com")
    assert "4 funds" in msg["Subject"] and "3 exact" in msg["Subject"]
    parts = {p.get_content_type() for p in msg.walk()}
    assert "text/plain" in parts and "text/html" in parts
    att = [p for p in msg.walk() if p.get_filename()]
    assert len(att) == 1 and att[0].get_filename().startswith("universe-")
    assert att[0].get_filename().endswith(".xlsx")
