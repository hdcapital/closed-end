#!/usr/bin/env python3
"""The universe as a spreadsheet a human would actually want to open.

Three sheets:

* **Summary** — what ran, what survived, and the discount picture by basis.
* **Universe** — every clean row, filterable, discounts coloured so the eye
  finds the wide ones without reading 390 numbers.
* **Dropped** — everything cleaning removed, with its reason. The audit trail
  travels with the data instead of living in a CI log.

Formatting rules worth stating: discounts are stored as fractions and shown by
Excel's percent format — never pre-multiplied, so the cell value and the CSV
agree. Money is whole currency units with thousands separators; magnitude
formatting belongs to the display layer, not the data.
"""

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .record import COLUMNS

INK = "1F2430"          # header band
ACCENT = "2563EB"       # links / highlights
MUTED = "6B7280"

_FMT = {
    "market_cap": "#,##0",
    "nta_total": "#,##0",
    "nta_per_share": "0.0000",
    "price": "0.0000",
    "discount": "+0.0%;-0.0%",
}
_WIDTH = {
    "code": 9, "isin": 15, "exchange": 10, "name": 38, "vehicle_type": 15,
    "sector": 26, "currency": 10, "market_cap": 16, "nta_total": 16,
    "nta_basis": 22, "nta_per_share": 12, "nta_unit": 15, "price": 10,
    "discount": 10, "discount_basis": 24, "nta_date": 12, "as_of": 12,
    "source": 16, "source_url": 40,
}

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor=INK)
_THIN = Border(bottom=Side(style="thin", color="E5E7EB"))


def _header(ws, labels, row=1):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def _sheet_of_rows(ws, columns, rows):
    _header(ws, columns)
    for r, row in enumerate(rows, start=2):
        for c, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(col))
            if col in _FMT and row.get(col) is not None:
                cell.number_format = _FMT[col]
            cell.border = _THIN
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _WIDTH.get(col, 12)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def build(rows, dropped, summary, path):
    """rows/dropped are dicts (CSV shape); summary is render-ready lines."""
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14
    ws["A1"] = "Closed-end universe"
    ws["A1"].font = Font(bold=True, size=16, color=INK)
    ws["A2"] = summary["subtitle"]
    ws["A2"].font = Font(size=10, color=MUTED)

    r = 4
    for label, value in summary["facts"]:
        ws.cell(row=r, column=1, value=label).font = Font(color=MUTED, size=10)
        ws.cell(row=r, column=2, value=value).font = Font(bold=True, size=10)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Discounts by basis").font = Font(bold=True, size=11)
    r += 1
    _header(ws, ["basis", "n", "p10", "median", "p90", "premiums"], row=r)
    for line in summary["by_basis"]:
        r += 1
        for c, v in enumerate(line, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (3, 4, 5):
                cell.number_format = "+0.0%;-0.0%"

    if summary.get("by_market"):
        r += 2
        ws.cell(row=r, column=1, value="Average discount by market").font = \
            Font(bold=True, size=11)
        r += 1
        _header(ws, ["market", "n", "average", "median"], row=r)
        for ex, n, avg, med in summary["by_market"]:
            r += 1
            for c, v in enumerate((ex, n, avg, med), start=1):
                cell = ws.cell(row=r, column=c, value=v)
                if c in (3, 4):
                    cell.number_format = "+0.0%;-0.0%"

    r += 2
    ws.cell(row=r, column=1, value="How to read the bases").font = Font(bold=True, size=11)
    for note in summary["notes"]:
        r += 1
        ws.cell(row=r, column=1, value=note).font = Font(size=9, color=MUTED)

    # --- Universe ---
    _sheet_of_rows(wb.create_sheet("Universe"), COLUMNS, rows)
    ws_u = wb["Universe"]
    disc_col = get_column_letter(COLUMNS.index("discount") + 1)
    if rows:
        # Red at a deep discount, white at par, green at a premium — the same
        # convention as every fund factsheet.
        ws_u.conditional_formatting.add(
            f"{disc_col}2:{disc_col}{len(rows) + 1}",
            ColorScaleRule(start_type="num", start_value=-0.5, start_color="F87171",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="num", end_value=0.15, end_color="4ADE80"))

    # --- Dropped ---
    _sheet_of_rows(wb.create_sheet("Dropped"),
                   ["code", "name", "reason", "market_cap", "nta_per_share"],
                   dropped)

    wb.save(path)
    return path
