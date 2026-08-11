#!/usr/bin/env python3
"""Tolerant spreadsheet reading.

The ASX monthly report and the LSE instrument list are human-facing
spreadsheets: title banners above the real header, footnotes below it, columns
renamed between editions ("ASX Code" -> "Code" -> "ASX code"), and the odd
merged cell. Hardcoding cell coordinates against files like these produces a
parser that works exactly once.

So: find the header row by looking for one that matches enough expected
patterns, then map logical field names to whichever column header matches.
When a required field can't be located, say so loudly — a silently mis-mapped
column is how a screen ends up ranking on the wrong number.
"""

import csv
import io
import re
from typing import Dict, List, Optional, Sequence


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def read_sheets(content: bytes, filename: str = "") -> Dict[str, List[list]]:
    """Return {sheet_name: rows}. Handles .xlsx/.xlsm via openpyxl and CSV.

    .xls (old binary) is deliberately unsupported: it needs xlrd, which no
    longer reads xls securely. If a source only offers .xls the caller should
    record that as an unsupported-format status rather than guess.
    """
    name = (filename or "").lower()
    head = content[:8]
    if head[:2] == b"PK":                       # zip container => xlsx/xlsm
        return _read_xlsx(content)
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise ValueError("legacy .xls (OLE2) format is not supported; need .xlsx or .csv")
    if name.endswith((".csv", ".txt")) or b"," in content[:4096] or b"\t" in content[:4096]:
        return {"csv": _read_csv(content)}
    raise ValueError(f"unrecognised spreadsheet format for {filename or 'document'}")


def _read_xlsx(content: bytes) -> Dict[str, List[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: Dict[str, List[list]] = {}
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            out[ws.title] = rows
    finally:
        wb.close()
    return out


def _read_csv(content: bytes) -> List[list]:
    text = content.decode("utf-8-sig", "replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(r) for r in csv.reader(io.StringIO(text), dialect)]


class ColumnMap:
    """Maps logical field -> column index, by fuzzy header match.

    A spec value is either a list of patterns, or a dict:

        {"match": [...], "not": [...]}

    The `not` list is what stops substring matching from being dangerous. The
    ASX report has a column headed "Prem/Disc % NTA (pre-tax) at N" sitting a
    few columns away from the real "NTA Price". A bare substring search for
    "nta pre tax" hits the *discount* column, and the resulting series looks
    like a fund whose NAV moved 100x in a month. Excluding any header
    containing "prem", "disc", "%", "return" or "change" from a *level* column
    makes that mistake unrepresentable rather than merely unlikely.
    """

    def __init__(self, header_row: List, spec: Dict[str, object]):
        self.header = [_norm(h) for h in header_row]
        self.raw_header = list(header_row)
        self.index: Dict[str, int] = {}
        self.missing: List[str] = []
        for field, rule in spec.items():
            if isinstance(rule, dict):
                patterns = rule.get("match", [])
                exclude = rule.get("not", [])
            else:
                patterns, exclude = rule, []
            idx = self._find(patterns, exclude)
            if idx is None:
                self.missing.append(field)
            else:
                self.index[field] = idx

    def _allowed(self, header: str, exclude: Sequence[str]) -> bool:
        return not any(_norm(x) in header for x in exclude if _norm(x))

    def _find(self, patterns: Sequence[str],
              exclude: Sequence[str] = ()) -> Optional[int]:
        # Exact normalised match first, then substring — in pattern order, so
        # the caller controls precedence ("nta price" before "nta").
        for pat in patterns:
            p = _norm(pat)
            for i, h in enumerate(self.header):
                if h == p and self._allowed(h, exclude):
                    return i
        for pat in patterns:
            p = _norm(pat)
            if not p:
                continue
            for i, h in enumerate(self.header):
                if p in h and self._allowed(h, exclude):
                    return i
        return None

    def get(self, row: List, field: str):
        i = self.index.get(field)
        if i is None or i >= len(row):
            return None
        return row[i]

    def has(self, field: str) -> bool:
        return field in self.index

    def header_for(self, field: str) -> str:
        """The raw header text a field mapped to.

        Callers use it to read the units the publisher declared rather than
        guessing them from the values: "MER (% p.a)" carrying 0.15 means
        0.15%, but a bare magnitude test would read it as 15%.
        """
        i = self.index.get(field)
        if i is None or i >= len(self.raw_header):
            return ""
        return str(self.raw_header[i] or "")


def describe(sheets: Dict[str, List[list]], max_rows: int = 6,
             max_cells: int = 14) -> str:
    """A compact picture of what a workbook actually contains.

    Attached to every parse failure. A parser that reports "no recognisable
    header" and nothing else forces a human to download the file to learn
    anything; one that shows the rows it rejected usually makes the fix
    obvious from the log alone.
    """
    out = []
    for name, rows in list(sheets.items())[:8]:
        out.append(f"sheet '{name}' ({len(rows)} rows)")
        shown = 0
        for i, row in enumerate(rows):
            cells = [str(c)[:28] for c in row[:max_cells] if c is not None and str(c).strip()]
            if not cells:
                continue
            out.append(f"    r{i}: {' | '.join(cells)}")
            shown += 1
            if shown >= max_rows:
                break
    return "\n".join(out)


def header_row_text(header_row: List) -> str:
    return " | ".join(str(c)[:30] for c in header_row if c is not None and str(c).strip())


def find_header(rows: List[list], required: Sequence[str],
                max_scan: int = 40) -> Optional[int]:
    """Index of the first row matching every `required` pattern.

    Requiring *all* patterns (rather than a count) keeps a banner row like
    "ASX Investment Products Monthly Report — Code of Practice" from being
    mistaken for the header just because it contains the word "code".
    """
    for i, row in enumerate(rows[:max_scan]):
        cells = [_norm(c) for c in row if c is not None]
        if not cells:
            continue
        if all(any(_norm(pat) in c for c in cells) for pat in required):
            return i
    return None


def data_rows(rows: List[list], header_idx: int, key_col: int):
    """Rows after the header that still look like data.

    Stops at the first run of blank keys, which is how these files separate the
    table from their footnotes ("Source: ASX", "* NTA is before tax", ...).
    """
    blanks = 0
    for row in rows[header_idx + 1:]:
        key = row[key_col] if key_col < len(row) else None
        if key is None or not str(key).strip():
            blanks += 1
            if blanks >= 3:
                break
            continue
        blanks = 0
        yield row
